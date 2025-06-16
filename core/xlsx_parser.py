import openpyxl
from openpyxl.utils import get_column_letter
import html

from core.schema import ContentType, BaseProperty, TextProperty, TableProperty, ImageProperty, FileBaseProperty
from core.parser import FileParser

from core.utils import num_tokens_from_string

class XlsxParser(FileParser):

    def __init__(self, file_path: str = None):
        super().__init__(file_path)


    def parse_sheet(self, sheet) -> TableProperty:
        """Parse a single sheet and return its properties."""
        html_content = f"<h2>{sheet.title}</h2>"
        html_content += '<table border="1">'
        
        # 获取合并单元格信息
        merged_ranges = sheet.merged_cells.ranges
        processed_cells = set()
        pure_text = ""
        for row in range(1, sheet.max_row + 1):
            html_content += "<tr>"
            for col in range(1, sheet.max_column + 1):
                if (row, col) in processed_cells:
                    continue

                cell = sheet.cell(row, col)

                # 检查是否在合并区域内
                rowspan, colspan = 1, 1
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        rowspan = merged_range.max_row - merged_range.min_row + 1
                        colspan = merged_range.max_col - merged_range.min_col + 1
                        # 标记合并区域内的所有单元格为已处理
                        for r in range(merged_range.min_row, merged_range.max_row + 1):
                            for c in range(merged_range.min_col, merged_range.max_col + 1):
                                processed_cells.add((r, c))
                        break
                
                # 生成HTML单元格
                cell_html = f'<td'
                if rowspan > 1:
                    cell_html += f' rowspan="{rowspan}"'
                if colspan > 1:
                    cell_html += f' colspan="{colspan}"'
                cell_html += f'>{html.escape(str(cell.value or ""))}</td>'
                html_content += cell_html
                pure_text += str(cell.value) + " "

            html_content += "</tr>"
        html_content += "</table>"

        table_property = TableProperty(name=f"Table {sheet.title}")
        table_property.html_content = html_content
        table_property.text_content = pure_text.strip()
        table_property.table_row_count = sheet.max_row
        table_property.table_column_count = sheet.max_column
        table_property.content_token_length = num_tokens_from_string(pure_text.strip())
        return table_property

    def parse(self, file_path: str) -> 'FileBaseProperty':
        workbook = openpyxl.load_workbook(file_path)
        self.property.page_count = len(workbook.sheetnames)
        page_index = 0
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            table_property = self.parse_sheet(sheet)
            self.property.total_text_length += len(table_property.text_content)
            self.property.total_token_length += table_property.content_token_length
            table_property.page_idx = page_index  # Assuming all sheets are on the first page
            table_property.name = f"Sheet: {sheet_name}"
            self.property.content_list.append(table_property)

            page_index += 1

        return self.property